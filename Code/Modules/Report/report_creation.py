import requests
import json
import os
import pandas as pd

from copy import deepcopy
from pathlib import Path
from glob import glob
from openpyxl import Workbook
from decouple import config

URL = "https://gemmoai--sound-api-fastapi-app-dev.modal.run"

# Get token key from .env file
token = config("REPORT_TOKEN")


def create_report(endpoint: str = "classify", client: str = "TestClient", **kwargs):
    # Define audio format
    formats = [
        ".wav",
        ".mp3",
        ".ogg",
        ".vorbis",
        ".amr-nb",
        ".amb",
        ".flac",
        ".sph",
        ".gsm",
        ".htk",
    ]

    # Get to audio directory
    audio_directory = os.path.join("Clients", client, "Audio")

    # Get file names
    file_names = []
    for ext in formats:
        file_names.extend(glob(audio_directory + f"/*{ext}"))

    # Process each file separately
    for file_name in file_names:
        # Call the API
        response = call_api(endpoint, file_name, **kwargs)

        # Catch issues with the API
        if response.status_code != 200:
            raise Exception(f"Not working: {response.text}")

        # Get metadata
        metadata = create_metadata_df(response, file_name, **kwargs)

        # Get the name of the excel file
        excel_name = os.path.join(
            "Clients", client, f"{client}_{endpoint}_{kwargs['model_name']}.xlsx"
        )

        if not os.path.exists(excel_name):
            # Create a new workbook
            wb = Workbook()

            # Create sheet
            wb.active.title = "Summary"

            # Save the workbook to the file
            wb.save(excel_name)

        # Write meta to Excel
        write_to_excel(file_name=excel_name, sheet_name="Summary", df=metadata)

        # Create the results df, the formatting depends on the endpoint
        if endpoint == "classify":
            df = create_results_df(response)
        elif endpoint == "window":
            df = create_results_df_window(response)

        # Write to Excel
        write_to_excel(file_name=excel_name, df=df, sheet_name=kwargs["model_name"])


def call_api(endpoint, file_names, **kwargs):
    # Create payload
    payload = kwargs

    # Create headers field with authorization token
    headers = {"Authorization": token}

    # Read the audio file(s)
    if endpoint == "classify":
        files = [("audio_files", (file_names, open(file_names, "rb")))]
    elif endpoint == "window":
        files = [("audio_file", (file_names, open(file_names, "rb")))]

    # Interrogate the API with a POST request
    response = requests.post(
        URL + "/" + endpoint, files=files, data=payload, headers=headers
    )

    return response


def create_results_df(response):
    # Create dictionary
    response_dictionary = json.loads(response.text)["classification"]

    # try:
    #     # Create dictionary
    #     response_dictionary = json.loads(response.text)["classification"]
    # except:
    #     # Create dictionary
    #     response_dictionary = json.loads(response.text)

    # Create an empty list to store the DataFrames
    dfs = []

    # Loop over the dictionary items and append each DataFrame to the list
    for key, value in response_dictionary.items():
        # Create a DataFrame from the tags list
        tags_df = pd.DataFrame(value["tags"], columns=["tag", "confidence"])

        # Create a hierarchical index with two levels
        idx = pd.MultiIndex.from_arrays(
            [[Path(key).stem] * len(tags_df), range(1, len(tags_df) + 1)],
            names=["filename", "tag_index"],
        )

        # Set the index of the DataFrame to the hierarchical index
        tags_df.index = idx

        # Append to list of dfs
        dfs.append(tags_df)

    return pd.concat(dfs)


def create_results_df_window(response):
    # Create dictionary
    response_dictionary = json.loads(response.text)

    # Extract the file name
    file_name = Path(response_dictionary["name"]).name

    # Select the part of the dictionary containing the windows
    windows = response_dictionary["windows"]

    # Init list where dfs are stored
    windows_df = []

    # Loop over each window found in there
    for window in windows:
        # Create a DataFrame from the tags list
        tags_df = pd.DataFrame(window["tags"], columns=["tag", "confidence"])

        # Save the start and end of the window
        tags_df["start"] = window["start"]
        tags_df["end"] = window["end"]

        # Add this window to the list of dfs
        windows_df.append(tags_df)

    # Create a hierarchical index with two levels
    idx = pd.MultiIndex.from_arrays(
        [[file_name] * len(windows_df), range(1, len(windows_df) + 1)],
        names=["filename", "window number"],
    )

    # Concatenate all the dfs
    df = pd.concat(windows_df)

    # Set the index
    df.index = idx

    # Reorganise the df
    df = df[["start", "end", "tag", "confidence"]]

    return df


def write_to_excel(file_name, df, sheet_name):
    """Write the findings onto an Excel file to be shared with client"""

    with pd.ExcelWriter(file_name, mode="a", if_sheet_exists="overlay") as writer:
        # Check if the sheet exists
        if sheet_name in writer.sheets:
            if sheet_name != "Summary":
                # Get the starting row for the new data
                startrow = writer.sheets[sheet_name].max_row
            else:
                if list(writer.sheets["Summary"].rows) == []:
                    startrow = 0
                else:
                    startrow = writer.sheets[sheet_name].max_row
        else:
            # The sheet doesn't exist, start at the first row
            startrow = 0

        if startrow > 0:
            # Write to file
            df.to_excel(
                writer,
                sheet_name=sheet_name,
                index=(not sheet_name == "Summary"),
                startrow=startrow,
                header=False,
            )
        else:
            df.to_excel(
                writer,
                sheet_name=sheet_name,
                index=(not sheet_name == "Summary"),
                startrow=startrow,
                header=True,
            )
    return


def create_metadata_df(response, file_names, **kwargs):
    """Create a df containing the metadata of the call"""

    # Copy dictionary
    args = deepcopy(kwargs)

    # Get model name
    model_name = args.pop("model_name")

    # Create dataframe with metadata
    metadata = pd.DataFrame(
        [
            [model_name, response.elapsed.total_seconds(), len(file_names)]
            + list(args.values())
        ],
        columns=["model", "time", "files"] + list(args.keys()),
    )

    return metadata
